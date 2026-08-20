import os, sqlite3, json, shutil, base64, io, re, secrets, hashlib, urllib.parse, threading, queue, time
from datetime import datetime, timedelta, timezone as _dt_timezone
from zoneinfo import ZoneInfo
from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for
from werkzeug.utils import secure_filename
from PIL import Image
import openpyxl

app = Flask(__name__)
app.secret_key = 'myhq-proposal-tool-secret'
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# Use /data volume on Railway (persistent), fall back to local for dev
_DATA_DIR = '/data' if os.path.isdir('/data') else BASE_DIR
DB_PATH = os.path.join(_DATA_DIR, 'proposals.db')
UPLOADS = os.path.join(BASE_DIR, 'uploads')
CENTRE_IMAGES = os.path.join(UPLOADS, 'centres')
PROPOSAL_FILES = os.path.join(UPLOADS, 'proposals')
TEMPLATE_FILES = os.path.join(UPLOADS, 'templates')

os.makedirs(CENTRE_IMAGES, exist_ok=True)
os.makedirs(PROPOSAL_FILES, exist_ok=True)
os.makedirs(TEMPLATE_FILES, exist_ok=True)

# ── Database ────────────────────────────────────────────────────────────────

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.executescript('''
        CREATE TABLE IF NOT EXISTS centres (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            address TEXT,
            city TEXT,
            about TEXT,
            space_type TEXT,
            brand TEXT,
            price_from INTEGER,
            price_unit TEXT DEFAULT 'MONTHLY',
            seat_type TEXT,
            open_hours TEXT DEFAULT '9:00 AM – 6:00 PM',
            amenities TEXT,
            transport TEXT,
            website TEXT,
            map_url TEXT,
            coordinates TEXT,
            why_recommend TEXT,
            source TEXT DEFAULT 'manual',
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS centre_images (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            centre_id INTEGER NOT NULL,
            filename TEXT NOT NULL,
            label TEXT,
            is_primary INTEGER DEFAULT 0,
            sort_order INTEGER DEFAULT 0,
            FOREIGN KEY (centre_id) REFERENCES centres(id) ON DELETE CASCADE
        );
        CREATE TABLE IF NOT EXISTS proposals (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT,
            template TEXT DEFAULT 'london',
            client_name TEXT,
            client_company TEXT,
            client_email TEXT,
            client_location TEXT,
            team_size TEXT,
            space_type TEXT,
            area_required TEXT,
            budget TEXT,
            duration TEXT,
            selected_centres TEXT DEFAULT '[]',
            manual_centres TEXT DEFAULT '[]',
            status TEXT DEFAULT 'draft',
            pptx_filename TEXT,
            pdf_filename TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        );
        CREATE TABLE IF NOT EXISTS share_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            token TEXT UNIQUE NOT NULL,
            label TEXT,
            centre_ids TEXT NOT NULL,
            centre_names TEXT NOT NULL,
            canonical_ids TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            created_by TEXT DEFAULT 'myHQ',
            client_email TEXT,
            client_phone TEXT
        );
        CREATE TABLE IF NOT EXISTS link_events (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            token TEXT NOT NULL,
            event_type TEXT NOT NULL,
            centre_id TEXT,
            centre_name TEXT,
            dwell_seconds INTEGER,
            ip_hash TEXT,
            user_agent TEXT,
            booking_date TEXT,
            booking_time TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        );
        ''')
    # migrate: add pdf_filename if missing
    with get_db() as conn:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(proposals)").fetchall()]
        if 'pdf_filename' not in cols:
            conn.execute("ALTER TABLE proposals ADD COLUMN pdf_filename TEXT")
    # migrate: add hotdesk_price (per day coworking price) if missing
    with get_db() as conn:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(centres)").fetchall()]
        if 'hotdesk_price' not in cols:
            conn.execute("ALTER TABLE centres ADD COLUMN hotdesk_price INTEGER")
        if 'has_coworking' not in cols:
            conn.execute("ALTER TABLE centres ADD COLUMN has_coworking INTEGER DEFAULT 0")
        if 'min_desks' not in cols:
            conn.execute("ALTER TABLE centres ADD COLUMN min_desks INTEGER")
        if 'hubble_id' not in cols:
            # Previously only added by the standalone import_hubble.py script — but
            # app.py code (matching, indexing, backup restore) depends on this column
            # existing on ANY database, so it belongs in the main schema migration too.
            conn.execute("ALTER TABLE centres ADD COLUMN hubble_id TEXT")
    # migrate: add new columns to share_links if missing
    with get_db() as conn:
        for col_def in ['client_email TEXT', 'client_phone TEXT', 'canonical_ids TEXT', 'personalised_message TEXT', 'recommended_ids TEXT']:
            col_name = col_def.split()[0]
            try:
                conn.execute(f'ALTER TABLE share_links ADD COLUMN {col_name} TEXT')
            except Exception:
                pass  # column already exists
        # Ensure indexes exist
        try:
            conn.execute('CREATE INDEX IF NOT EXISTS idx_share_links_canonical ON share_links(canonical_ids)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_link_events_token_type ON link_events(token, event_type)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_link_events_token ON link_events(token)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_centres_hubble_id ON centres(hubble_id)')
            conn.execute('CREATE INDEX IF NOT EXISTS idx_centre_images_centre_id ON centre_images(centre_id)')
        except Exception:
            pass
    # migrate: add booking columns to link_events if missing
    with get_db() as conn:
        cols = [r[1] for r in conn.execute('PRAGMA table_info(link_events)').fetchall()]
        if 'booking_date' not in cols:
            conn.execute('ALTER TABLE link_events ADD COLUMN booking_date TEXT')
            conn.execute('ALTER TABLE link_events ADD COLUMN booking_time TEXT')
    # migrate: add is_test flag to share_links — lets test/dev links be excluded from analytics
    with get_db() as conn:
        cols = [r[1] for r in conn.execute('PRAGMA table_info(share_links)').fetchall()]
        if 'is_test' not in cols:
            conn.execute('ALTER TABLE share_links ADD COLUMN is_test INTEGER DEFAULT 0')
    # migrate: 'open' and 'dwell' events were sometimes logged twice for one real
    # visit (browser/beacon quirk) — clean up existing duplicates, then enforce
    # one-per-second-per-visitor at the DB level so it can't happen again.
    with get_db() as conn:
        conn.execute("""
            DELETE FROM link_events
            WHERE event_type IN ('open','dwell')
            AND id NOT IN (
                SELECT MIN(id) FROM link_events
                WHERE event_type IN ('open','dwell')
                GROUP BY token, ip_hash, event_type, created_at
            )
        """)
        conn.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_dedup_open
            ON link_events(token, ip_hash, created_at) WHERE event_type='open'
        """)
        conn.execute("""
            CREATE UNIQUE INDEX IF NOT EXISTS idx_dedup_dwell
            ON link_events(token, ip_hash, created_at) WHERE event_type='dwell'
        """)

# SSE event broadcast for real-time dashboard
_sse_queues = []
_sse_lock = threading.Lock()

def _push_sse(data):
    with _sse_lock:
        for q in list(_sse_queues):
            try: q.put_nowait(data)
            except: pass

init_db()

# ── Helpers ─────────────────────────────────────────────────────────────────

ALLOWED = {'png','jpg','jpeg','gif','webp'}

def allowed(filename):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in ALLOWED

def save_image(file_obj, folder, name_prefix='img'):
    filename = secure_filename(file_obj.filename)
    ext = filename.rsplit('.',1)[-1].lower() if '.' in filename else 'jpg'
    out_name = f"{name_prefix}_{os.urandom(6).hex()}.{ext}"
    out_path = os.path.join(folder, out_name)
    img = Image.open(file_obj)
    img = img.convert('RGB')
    img.save(out_path, quality=90)
    return out_name

def save_image_bytes(data_bytes, folder, name_prefix='img'):
    out_name = f"{name_prefix}_{os.urandom(6).hex()}.jpg"
    out_path = os.path.join(folder, out_name)
    img = Image.open(io.BytesIO(data_bytes)).convert('RGB')
    img.save(out_path, quality=90)
    return out_name

def centre_image_dir(centre_id):
    d = os.path.join(CENTRE_IMAGES, str(centre_id))
    os.makedirs(d, exist_ok=True)
    return d

# ── Excel import ────────────────────────────────────────────────────────────

def _row_get(row, idx):
    """Safe positional read — Excel drops trailing blank cells, so rows are
    often shorter than the header implies."""
    return row[idx] if row and idx is not None and idx < len(row) else None

def import_excel(filepath):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['workspace']
    headers = [c for c in next(ws.iter_rows(values_only=True))]
    h = {v: i for i, v in enumerate(headers) if v}

    pricing = {}
    for sheet in ['dedicated_price_plans', 'flexi_price_plans', 'vo_price_plans']:
        if sheet not in wb.sheetnames:
            continue
        pws = wb[sheet]
        ph = None
        for row in pws.iter_rows(values_only=True):
            if ph is None:
                ph = {v: i for i, v in enumerate(row) if v}
                continue
            wid = _row_get(row, ph.get('workspace_identifier', 0))
            if not wid:
                continue
            amt = _row_get(row, ph.get('amount', ph.get('pricePerSeat', 6)))
            unit = _row_get(row, ph.get('paymentCycle', ph.get('unit', 3)))  # paymentCycle = MONTHLY/DAILY
            stype = _row_get(row, ph.get('seatType', ph.get('type', 4)))
            if wid not in pricing and amt:
                pricing[wid] = {'amount': amt, 'unit': unit or 'MONTHLY', 'seat_type': stype}

    amenities_map = {}
    for sheet in ['dedicated_amenities', 'flexi_amenities', 'vo_amenities', 'meeting_room_amenities']:
        if sheet not in wb.sheetnames:
            continue
        aws = wb[sheet]
        ah = None
        for row in aws.iter_rows(values_only=True):
            if ah is None:
                ah = {v: i for i, v in enumerate(row) if v}
                continue
            wid = _row_get(row, ah.get('workspace_identifier', 0))
            slug = _row_get(row, ah.get('amenity_slug', 1))
            if wid and slug:
                amenities_map.setdefault(wid, [])
                if slug not in amenities_map[wid]:
                    amenities_map[wid].append(slug)

    added = 0
    updated = 0
    skipped = 0
    ambiguous = []
    with get_db() as conn:
        # Match existing centres by address first (names in operator uploads are
        # often just the brand, e.g. "WeWork", so name alone can't disambiguate).
        centres = conn.execute('SELECT id, name, address FROM centres').fetchall()
        addr_index = {}
        for c in centres:
            if c['address']:
                addr_index.setdefault(_norm_address(c['address']), []).append(c)
        name_index = {}
        for c in centres:
            name_index.setdefault(_norm_text(c['name']), []).append(c)

        for row in ws.iter_rows(values_only=True):
            first = _row_get(row, 0)
            if first == 'workspace_identifier' or first is None:
                continue
            wid = _row_get(row, h.get('workspace_identifier', 0))
            name = _row_get(row, h.get('name', 1))
            addr = _row_get(row, h.get('address', 4))
            if not name and not addr:
                skipped += 1
                continue

            match = None
            if addr:
                match, is_ambiguous = _find_address_match(addr, centres, addr_index)
                if is_ambiguous:
                    ambiguous.append({'name': name, 'address': addr})
                    continue
            if not match and name:
                candidates = name_index.get(_norm_text(name), [])
                if len(candidates) == 1:
                    match = candidates[0]

            city = _row_get(row, h.get('city_slug', 5))
            about = _row_get(row, h.get('about', 8))
            stype = _row_get(row, h.get('spaceType', 9))
            brand = _row_get(row, h.get('workspaceBrand_slug', 11))
            transport_raw = _row_get(row, h.get('directions_connectivityDetails', 12))
            bus = _row_get(row, h.get('directions_nearestBus', 14))
            coords = _row_get(row, h.get('loc_coordinates', 6))
            mapurl = _row_get(row, h.get('mapurl', 7))

            transport_lines = []
            if transport_raw:
                for part in str(transport_raw).split(','):
                    part = part.strip()
                    segments = part.split(':')
                    if len(segments) >= 2:
                        transport_lines.append(segments[1].strip())
            if bus:
                transport_lines.append(f'Bus: {bus}')
            transport = ', '.join(transport_lines) if transport_lines else ''

            price_info = pricing.get(wid, {})
            amenities_json = json.dumps(amenities_map.get(wid, []))

            if match:
                # Deliberately leave `name` untouched — the fuller name already in the
                # DB (e.g. "WeWork - 10 York Road") is better than the generic upload name.
                conn.execute('''UPDATE centres SET
                    address=?, city=?, about=?, space_type=?, brand=?, price_from=?,
                    price_unit=?, seat_type=?, amenities=?, transport=?, map_url=?, coordinates=?
                    WHERE id=?''',
                    (addr, city, about, stype, brand,
                     price_info.get('amount'), price_info.get('unit', 'MONTHLY'),
                     price_info.get('seat_type'), amenities_json, transport, mapurl, coords,
                     match['id']))
                updated += 1
            else:
                conn.execute('''INSERT INTO centres
                    (name,address,city,about,space_type,brand,price_from,price_unit,seat_type,amenities,transport,map_url,coordinates,source)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
                    (name, addr, city, about, stype, brand,
                     price_info.get('amount'), price_info.get('unit','MONTHLY'),
                     price_info.get('seat_type'), amenities_json,
                     transport, mapurl, coords, 'excel'))
                added += 1
    wb.close()
    return added, updated, skipped, ambiguous

# ── Address/name matching helpers (used by import_excel to detect existing centres) ──

def _norm_text(s):
    return re.sub(r'\s+', ' ', (s or '').strip().lower())

def _norm_address(s):
    # Strip ALL punctuation (including commas) — operator files often punctuate
    # the same address differently (e.g. "125 Kingsway London, WC2B" vs
    # "125 Kingsway, London WC2B"), so comma position must not affect matching.
    s = re.sub(r'[^\w\s]', '', _norm_text(s))
    return re.sub(r'\s+', ' ', s).strip()

def _find_address_match(addr, centres, addr_index, min_jaccard=0.6, tie_margin=0.15):
    """Find the centre whose address matches `addr`. Exact match (after
    normalization) wins outright. Otherwise falls back to word-overlap
    (Jaccard) similarity, since operators often insert extra words into an
    address ("Waterloo", "138 Holborn", "London") that breaks plain substring
    matching. Returns (matched_centre_or_None, is_ambiguous)."""
    norm_addr = _norm_address(addr)
    exact = addr_index.get(norm_addr, [])
    if len(exact) == 1:
        return exact[0], False
    if len(exact) > 1:
        return None, True

    addr_tokens = set(norm_addr.split())
    if not addr_tokens:
        return None, False
    scored = []
    for c in centres:
        if not c['address']:
            continue
        c_tokens = set(_norm_address(c['address']).split())
        if not c_tokens:
            continue
        union = addr_tokens | c_tokens
        jaccard = len(addr_tokens & c_tokens) / len(union) if union else 0
        if jaccard >= min_jaccard:
            scored.append((jaccard, c))
    if not scored:
        return None, False
    scored.sort(key=lambda x: -x[0])
    if len(scored) == 1 or scored[0][0] - scored[1][0] > tie_margin:
        return scored[0][1], False
    return None, True  # top two candidates too close to call — don't guess

# ── Photo import from Drive folder links ─────────────────────────────────────
# Operator workbooks (e.g. WeWork's) reference a public Google Drive folder per
# centre via image_folder_url instead of direct image files. This downloads
# each folder's images and attaches them to the matching centre.

def _download_drive_folder_images(folder_url, dest_dir, name_prefix):
    import gdown, tempfile
    tmp_dir = tempfile.mkdtemp(prefix='drive_dl_')
    saved = []
    try:
        files = gdown.download_folder(url=folder_url, output=tmp_dir, quiet=True, use_cookies=False) or []
        for fp in files:
            try:
                with open(fp, 'rb') as fh:
                    raw = fh.read()
                saved.append(save_image_bytes(raw, dest_dir, name_prefix))
            except Exception:
                continue
    finally:
        shutil.rmtree(tmp_dir, ignore_errors=True)
    return saved

def import_photos_from_excel(filepath, overwrite=False):
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb['workspace']
    rows = list(ws.iter_rows(values_only=True))
    wb.close()  # release the file handle now — read_only workbooks keep it open otherwise
    headers = rows[0]
    h = {v: i for i, v in enumerate(headers) if v}

    processed, skipped_has_images, no_folder_url, errors = [], [], [], []
    with get_db() as conn:
        centres = conn.execute('SELECT id, name, address FROM centres').fetchall()
        addr_index = {}
        for c in centres:
            if c['address']:
                addr_index.setdefault(_norm_address(c['address']), []).append(c)

        for row in rows[1:]:
            first = _row_get(row, 0)
            if first == 'workspace_identifier' or first is None:
                continue
            name = _row_get(row, h.get('name', 1))
            addr = _row_get(row, h.get('address', 4))
            folder_url = _row_get(row, h.get('image_folder_url')) or _row_get(row, h.get('ext_image_folder_url'))
            if not addr or not folder_url:
                no_folder_url.append({'name': name, 'address': addr})
                continue

            centre, is_ambiguous = _find_address_match(addr, centres, addr_index)
            if not centre:
                errors.append({'name': name, 'address': addr,
                               'reason': 'ambiguous match' if is_ambiguous else 'no matching centre'})
                continue
            cid = centre['id']

            existing = conn.execute('SELECT COUNT(*) FROM centre_images WHERE centre_id=?', (cid,)).fetchone()[0]
            if existing and not overwrite:
                skipped_has_images.append({'id': cid, 'name': centre['name']})
                continue

            try:
                saved = _download_drive_folder_images(folder_url, centre_image_dir(cid), f'centre_{cid}_drive')
            except Exception as e:
                errors.append({'name': centre['name'], 'reason': str(e)})
                continue
            if not saved:
                errors.append({'name': centre['name'], 'reason': 'no images found in folder'})
                continue

            for i, fname in enumerate(saved):
                is_primary = 1 if (existing == 0 and i == 0) else 0
                conn.execute('INSERT INTO centre_images (centre_id, filename, is_primary, sort_order) VALUES (?,?,?,?)',
                             (cid, fname, is_primary, existing + i))
            processed.append({'id': cid, 'name': centre['name'], 'added': len(saved)})

    return {'processed': processed, 'skipped_has_images': skipped_has_images,
            'no_folder_url': no_folder_url, 'errors': errors}

# ── Routes: Centres ─────────────────────────────────────────────────────────

@app.route('/')
def index():
    with get_db() as conn:
        centres = conn.execute('SELECT id,name,city,space_type,brand,price_from,price_unit FROM centres ORDER BY name').fetchall()
        proposals = conn.execute('SELECT id,title,client_name,client_company,status,created_at FROM proposals ORDER BY created_at DESC LIMIT 5').fetchall()
        centre_count = conn.execute('SELECT COUNT(*) FROM centres').fetchone()[0]
    return render_template('index.html', centres=centres, proposals=proposals, centre_count=centre_count)

@app.route('/centres')
def centres_list():
    q = request.args.get('q','').strip()
    city = request.args.get('city','')
    space_type = request.args.get('space_type','')
    with get_db() as conn:
        query = 'SELECT c.*, (SELECT filename FROM centre_images WHERE centre_id=c.id AND is_primary=1 LIMIT 1) as primary_image FROM centres c WHERE 1=1'
        params = []
        if q:
            # Match each word separately (against name OR address) rather than the
            # whole phrase as one substring, so "wework moor place" finds a centre
            # whose brand is in the name and the rest of the words are in the address.
            for token in q.split():
                query += ' AND (c.name LIKE ? OR c.address LIKE ?)'
                params += [f'%{token}%', f'%{token}%']
        if city:
            query += ' AND c.city=?'
            params.append(city)
        if space_type:
            query += ' AND c.space_type=?'
            params.append(space_type)
        query += ' ORDER BY c.name'
        centres = conn.execute(query, params).fetchall()
        cities = [r[0] for r in conn.execute('SELECT DISTINCT city FROM centres WHERE city IS NOT NULL ORDER BY city').fetchall()]
        space_types = [r[0] for r in conn.execute('SELECT DISTINCT space_type FROM centres WHERE space_type IS NOT NULL ORDER BY space_type').fetchall()]
    return render_template('centres.html', centres=centres, cities=cities, space_types=space_types, q=q, city=city, space_type=space_type)

@app.route('/centres/<int:cid>')
def centre_detail(cid):
    with get_db() as conn:
        centre = conn.execute('SELECT * FROM centres WHERE id=?', (cid,)).fetchone()
        if not centre:
            return 'Not found', 404
        images = conn.execute('SELECT * FROM centre_images WHERE centre_id=? ORDER BY is_primary DESC, sort_order', (cid,)).fetchall()
    return render_template('centre_detail.html', centre=centre, images=images)

@app.route('/centres/add', methods=['POST'])
def centre_add():
    data = request.json or request.form
    amenities = data.get('amenities', '[]')
    if isinstance(amenities, str):
        try:
            json.loads(amenities)
        except:
            amenities = json.dumps([a.strip() for a in amenities.split(',') if a.strip()])
    with get_db() as conn:
        cur = conn.execute('''INSERT INTO centres
            (name,address,city,about,space_type,brand,price_from,price_unit,open_hours,amenities,transport,website,map_url,why_recommend,source)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)''',
            (data.get('name'), data.get('address'), data.get('city'),
             data.get('about'), data.get('space_type'), data.get('brand'),
             data.get('price_from') or None, data.get('price_unit','MONTHLY'),
             data.get('open_hours','9:00 AM – 6:00 PM'),
             amenities, data.get('transport'), data.get('website'),
             data.get('map_url'), data.get('why_recommend'), 'manual'))
        return jsonify({'id': cur.lastrowid, 'ok': True})

@app.route('/centres/<int:cid>/update', methods=['POST'])
def centre_update(cid):
    data = request.json or request.form
    fields = ['name','address','city','about','space_type','brand','price_from','price_unit',
              'open_hours','amenities','transport','website','map_url','why_recommend']
    sets = ', '.join(f'{f}=?' for f in fields if f in data)
    vals = [data[f] for f in fields if f in data] + [cid]
    if sets:
        with get_db() as conn:
            conn.execute(f'UPDATE centres SET {sets} WHERE id=?', vals)
    return jsonify({'ok': True})

@app.route('/centres/<int:cid>/delete', methods=['POST','DELETE'])
def centre_delete(cid):
    img_dir = os.path.join(CENTRE_IMAGES, str(cid))
    if os.path.isdir(img_dir):
        shutil.rmtree(img_dir)
    with get_db() as conn:
        conn.execute('DELETE FROM centre_images WHERE centre_id=?', (cid,))
        conn.execute('DELETE FROM centres WHERE id=?', (cid,))
    return jsonify({'ok': True})

@app.route('/centres/<int:cid>/upload-image', methods=['POST'])
def centre_upload_image(cid):
    img_dir = centre_image_dir(cid)
    filenames = []

    # Handle file upload
    for f in request.files.getlist('images'):
        if f and allowed(f.filename):
            name = save_image(f, img_dir, f'centre_{cid}')
            filenames.append(name)

    # Handle base64 paste
    for paste_data in request.form.getlist('paste_data'):
        if paste_data.startswith('data:image'):
            header, b64 = paste_data.split(',', 1)
            raw = base64.b64decode(b64)
            name = save_image_bytes(raw, img_dir, f'centre_{cid}_paste')
            filenames.append(name)

    with get_db() as conn:
        existing = conn.execute('SELECT COUNT(*) FROM centre_images WHERE centre_id=?', (cid,)).fetchone()[0]
        for i, fname in enumerate(filenames):
            is_primary = 1 if (existing == 0 and i == 0) else 0
            conn.execute('INSERT INTO centre_images (centre_id, filename, is_primary, sort_order) VALUES (?,?,?,?)',
                         (cid, fname, is_primary, existing + i))

    images = []
    with get_db() as conn:
        rows = conn.execute('SELECT * FROM centre_images WHERE centre_id=? ORDER BY is_primary DESC, sort_order', (cid,)).fetchall()
        for r in rows:
            images.append({'id': r['id'], 'filename': r['filename'], 'is_primary': r['is_primary']})
    return jsonify({'ok': True, 'images': images})

@app.route('/centres/<int:cid>/set-primary-image/<int:img_id>', methods=['POST'])
def set_primary_image(cid, img_id):
    with get_db() as conn:
        conn.execute('UPDATE centre_images SET is_primary=0 WHERE centre_id=?', (cid,))
        conn.execute('UPDATE centre_images SET is_primary=1 WHERE id=? AND centre_id=?', (img_id, cid))
    return jsonify({'ok': True})

@app.route('/centres/<int:cid>/delete-image/<int:img_id>', methods=['POST','DELETE'])
def delete_image(cid, img_id):
    with get_db() as conn:
        row = conn.execute('SELECT filename FROM centre_images WHERE id=? AND centre_id=?', (img_id, cid)).fetchone()
        if row:
            fpath = os.path.join(CENTRE_IMAGES, str(cid), row['filename'])
            if os.path.exists(fpath):
                os.remove(fpath)
            conn.execute('DELETE FROM centre_images WHERE id=?', (img_id,))
    return jsonify({'ok': True})

@app.route('/centres/<int:cid>/fetch-photos', methods=['POST'])
def centre_fetch_photos(cid):
    import urllib.request, urllib.parse
    data = request.json or {}
    api_key = data.get('api_key') or os.environ.get('GOOGLE_MAPS_API_KEY', '')
    if not api_key:
        return jsonify({'error': 'No API key provided'}), 400

    with get_db() as conn:
        centre = conn.execute('SELECT name, address FROM centres WHERE id=?', (cid,)).fetchone()
    if not centre:
        return jsonify({'error': 'Centre not found'}), 404

    name = centre['name'] or ''
    address = centre['address'] or ''
    query = f"{name} {address}".strip()

    # 1. Find place
    search_url = (
        'https://maps.googleapis.com/maps/api/place/findplacefromtext/json?'
        + urllib.parse.urlencode({'input': query, 'inputtype': 'textquery', 'fields': 'photos', 'key': api_key})
    )
    try:
        with urllib.request.urlopen(search_url, timeout=10) as r:
            result = json.loads(r.read().decode())
    except Exception as e:
        return jsonify({'error': f'Places API error: {e}'}), 500

    candidates = result.get('candidates', [])
    if not candidates:
        return jsonify({'ok': True, 'added': 0, 'note': 'No place found'})

    photos = candidates[0].get('photos', [])[:4]
    if not photos:
        return jsonify({'ok': True, 'added': 0, 'note': 'No photos for this place'})

    img_dir = centre_image_dir(cid)
    added = 0
    with get_db() as conn:
        existing_count = conn.execute('SELECT COUNT(*) FROM centre_images WHERE centre_id=?', (cid,)).fetchone()[0]
        for i, photo in enumerate(photos):
            ref = photo.get('photo_reference')
            if not ref:
                continue
            photo_url = (
                'https://maps.googleapis.com/maps/api/place/photo?'
                + urllib.parse.urlencode({'maxwidth': 800, 'photo_reference': ref, 'key': api_key})
            )
            try:
                req = urllib.request.Request(photo_url, headers={'User-Agent': 'myHQ-proposal-tool'})
                with urllib.request.urlopen(req, timeout=15) as r:
                    raw = r.read()
                out_name = save_image_bytes(raw, img_dir, f'centre_{cid}_gp')
                is_primary = 1 if (existing_count == 0 and i == 0) else 0
                conn.execute(
                    'INSERT INTO centre_images (centre_id, filename, is_primary, sort_order) VALUES (?,?,?,?)',
                    (cid, out_name, is_primary, existing_count + i)
                )
                added += 1
            except Exception:
                continue
    return jsonify({'ok': True, 'added': added})

@app.route('/centres/<int:cid>/import-url', methods=['POST'])
def centre_import_url(cid):
    """Fetch an image from a URL and save it as a centre image."""
    import urllib.request, ssl
    data = request.json or {}
    url = (data.get('url') or '').strip()
    if not url:
        return jsonify({'error': 'No URL provided'}), 400

    with get_db() as conn:
        if not conn.execute('SELECT 1 FROM centres WHERE id=?', (cid,)).fetchone():
            return jsonify({'error': 'Centre not found'}), 404

    ctx = ssl._create_unverified_context()
    try:
        req = urllib.request.Request(url, headers={
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
            'Referer': url,
        })
        with urllib.request.urlopen(req, timeout=15, context=ctx) as r:
            raw = r.read()
    except Exception as e:
        return jsonify({'error': f'Failed to fetch URL: {e}'}), 400

    if len(raw) < 1000:
        return jsonify({'error': 'Response too small — not a valid image'}), 400

    img_dir = centre_image_dir(cid)
    try:
        fname = save_image_bytes(raw, img_dir, f'centre_{cid}_url')
    except Exception as e:
        return jsonify({'error': f'Could not save image: {e}'}), 500

    with get_db() as conn:
        existing = conn.execute('SELECT COUNT(*) FROM centre_images WHERE centre_id=?', (cid,)).fetchone()[0]
        is_primary = 1 if existing == 0 else 0
        conn.execute(
            'INSERT INTO centre_images (centre_id, filename, is_primary, sort_order) VALUES (?,?,?,?)',
            (cid, fname, is_primary, existing)
        )
        rows = conn.execute(
            'SELECT * FROM centre_images WHERE centre_id=? ORDER BY is_primary DESC, sort_order', (cid,)
        ).fetchall()
        images = [{'id': r['id'], 'filename': r['filename'], 'is_primary': r['is_primary']} for r in rows]

    return jsonify({'ok': True, 'images': images})


@app.route('/centres/import-excel', methods=['POST'])
def import_excel_route():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file'}), 400
    tmp = os.path.join(UPLOADS, 'tmp_import.xlsx')
    f.save(tmp)
    try:
        added, updated, skipped, ambiguous = import_excel(tmp)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except OSError:
            pass  # a transient lock on the temp file must not hide a successful import
    return jsonify({'ok': True, 'added': added, 'updated': updated, 'skipped': skipped, 'ambiguous': ambiguous})

@app.route('/centres/import-photos-from-excel', methods=['POST'])
def import_photos_from_excel_route():
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file'}), 400
    tmp = os.path.join(UPLOADS, 'tmp_photos_import.xlsx')
    f.save(tmp)
    try:
        result = import_photos_from_excel(tmp)
    except Exception as e:
        return jsonify({'error': str(e)}), 500
    finally:
        try:
            if os.path.exists(tmp):
                os.remove(tmp)
        except OSError:
            pass  # a transient lock on the temp file must not hide a successful import
    return jsonify({'ok': True, **result})

@app.route('/centres/add-page')
def centres_add_page():
    return redirect(url_for('centres_list'))

@app.route('/centres/api/search')
def api_centre_search():
    q = request.args.get('q','').strip()
    with get_db() as conn:
        rows = conn.execute(
            '''SELECT c.id, c.name, c.address, c.city, c.space_type, c.price_from, c.price_unit,
               c.hotdesk_price,
               (SELECT filename FROM centre_images WHERE centre_id=c.id AND is_primary=1 LIMIT 1) as primary_image
               FROM centres c
               WHERE c.name LIKE ? OR c.address LIKE ?
               ORDER BY c.name LIMIT 30''',
            (f'%{q}%', f'%{q}%')).fetchall()
    return jsonify([dict(r) for r in rows])

@app.route('/centres/api/<int:cid>')
def api_centre_detail(cid):
    with get_db() as conn:
        c = conn.execute('SELECT * FROM centres WHERE id=?', (cid,)).fetchone()
        if not c:
            return jsonify({'error': 'Not found'}), 404
        imgs = conn.execute('SELECT id, filename, label, is_primary, sort_order FROM centre_images WHERE centre_id=? ORDER BY is_primary DESC, sort_order', (cid,)).fetchall()
    centre = dict(c)
    img_list = [dict(i) for i in imgs]
    # images: list of URLs (for builder display)
    def _img_url(cid, filename):
        if filename.startswith('http'):
            return filename  # Hubble CDN URL — use directly
        return f'/centre-image/{cid}/{filename}'

    centre['images'] = [_img_url(cid, i['filename']) for i in img_list]
    # images_meta: full records (for image editor modal)
    centre['images_meta'] = img_list
    return jsonify(centre)

# ── Routes: Image serving ────────────────────────────────────────────────────

@app.route('/centre-image/<int:cid>/<path:filename>')
def serve_centre_image(cid, filename):
    # If filename is a full URL (shouldn't normally happen but handle gracefully)
    if filename.startswith('http'):
        from flask import redirect
        return redirect(filename)
    img_dir = os.path.join(CENTRE_IMAGES, str(cid))
    return send_file(os.path.join(img_dir, filename))

# ── Routes: Image proxy (for manual/non-DB centres) ─────────────────────────

@app.route('/api/fetch-image-b64', methods=['POST'])
def api_fetch_image_b64():
    """Fetch an image from a URL and return it as a base64 data URI."""
    data = request.json or {}
    url = (data.get('url') or '').strip()
    if not url or not url.startswith('http'):
        return jsonify({'error': 'Invalid URL'}), 400
    try:
        import urllib.request, base64 as _b64
        req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
        with urllib.request.urlopen(req, timeout=10) as resp:
            content_type = resp.headers.get('Content-Type', 'image/jpeg').split(';')[0].strip()
            raw = resp.read()
        b64 = f'data:{content_type};base64,{_b64.b64encode(raw).decode()}'
        return jsonify({'b64': b64})
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ── Routes: Map data API ─────────────────────────────────────────────────────

@app.route('/api/map-data')
def api_map_data():
    base = _base_url()
    with get_db() as conn:
        centres = conn.execute('SELECT * FROM centres ORDER BY name').fetchall()
        result = []
        for c in centres:
            cid = c['id']
            imgs = conn.execute(
                'SELECT filename FROM centre_images WHERE centre_id=? ORDER BY is_primary DESC, sort_order',
                (cid,)).fetchall()
            photos = [
                row["filename"] if row["filename"].startswith('http')
                else f'{base}/centre-image/{cid}/{row["filename"]}'
                for row in imgs
            ]
            entry = dict(c)
            entry['photos'] = photos
            result.append(entry)
    resp = jsonify(result)
    resp.headers['Access-Control-Allow-Origin'] = '*'
    return resp

# ── Routes: Proposals ───────────────────────────────────────────────────────

@app.route('/projects')
def projects_list():
    with get_db() as conn:
        proposals = conn.execute('SELECT * FROM proposals ORDER BY created_at DESC').fetchall()
    # Group by client_company (fall back to client_name then 'Unknown')
    groups = {}
    for p in proposals:
        key = p['client_company'] or p['client_name'] or 'Unknown'
        groups.setdefault(key, []).append(p)
    # Sort groups: most recently updated first
    sorted_groups = sorted(groups.items(), key=lambda kv: kv[1][0]['created_at'], reverse=True)
    return render_template('projects.html', groups=sorted_groups)

@app.route('/proposals')
def proposals_list():
    with get_db() as conn:
        proposals = conn.execute('SELECT * FROM proposals ORDER BY created_at DESC').fetchall()
    return render_template('proposals.html', proposals=proposals)

@app.route('/proposals/new', methods=['GET','POST'])
def proposal_new():
    if request.method == 'POST':
        data = request.json
        with get_db() as conn:
            cur = conn.execute('''INSERT INTO proposals
                (title,template,client_name,client_company,client_email,client_location,
                 team_size,space_type,area_required,budget,duration)
                VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (data.get('title','New Proposal'),
                 data.get('template','london'),
                 data.get('client_name'), data.get('client_company'),
                 data.get('client_email'), data.get('client_location'),
                 data.get('team_size'), data.get('space_type'),
                 data.get('area_required'), data.get('budget'),
                 data.get('duration')))
            pid = cur.lastrowid
        return jsonify({'ok': True, 'id': pid})
    # Handle map_ws param (base64 JSON workspace array from London map)
    preselected_ids = []
    map_workspaces = []
    map_ws_b64 = request.args.get('map_ws', '').strip()
    if map_ws_b64:
        try:
            import base64
            decoded = base64.b64decode(map_ws_b64).decode('utf-8')
            spaces = json.loads(decoded)
            if isinstance(spaces, list):
                for sp in spaces:
                    if not isinstance(sp, dict) or not sp.get('name'):
                        continue
                    row = None
                    with get_db() as conn:
                        # Try hubble_id first (most reliable)
                        hid = sp.get('hubble_id', '').strip()
                        if hid:
                            row = conn.execute(
                                'SELECT id FROM centres WHERE hubble_id=? LIMIT 1', (hid,)
                            ).fetchone()
                        # Fall back to exact name match
                        if not row:
                            row = conn.execute(
                                'SELECT id FROM centres WHERE LOWER(name)=LOWER(?) LIMIT 1',
                                (sp['name'].strip(),)
                            ).fetchone()
                        # Fall back to partial name match
                        if not row:
                            row = conn.execute(
                                'SELECT id FROM centres WHERE LOWER(name) LIKE LOWER(?) LIMIT 1',
                                (f'%{sp["name"]}%',)
                            ).fetchone()
                    if row:
                        preselected_ids.append(row['id'])
                    else:
                        # Workspace not in DB — pass as manual entry with Hubble image URLs
                        if isinstance(sp.get('amenities'), list):
                            sp['amenities'] = json.dumps(sp['amenities'])
                        # Store image URLs so generation.py can download them
                        if not sp.get('images') and sp.get('image_urls'):
                            sp['images'] = sp['image_urls']
                        map_workspaces.append(sp)
        except Exception:
            pass
    # Legacy fallback: map_names param
    elif request.args.get('map_names', '').strip():
        map_names = request.args.get('map_names', '').strip()
        names = [n.strip() for n in map_names.split(',') if n.strip()]
        with get_db() as conn:
            for name in names:
                row = conn.execute(
                    'SELECT id FROM centres WHERE LOWER(name) LIKE LOWER(?) LIMIT 1',
                    (f'%{name}%',)
                ).fetchone()
                if row:
                    preselected_ids.append(row['id'])
    map_space_type = request.args.get('space_type', '').strip()
    return render_template('builder.html', step=1, proposal=None,
                           preselected_ids=preselected_ids,
                           map_workspaces=map_workspaces,
                           map_space_type=map_space_type)

@app.route('/proposals/<int:pid>')
def proposal_builder(pid):
    with get_db() as conn:
        proposal = conn.execute('SELECT * FROM proposals WHERE id=?', (pid,)).fetchone()
        if not proposal:
            return 'Not found', 404
    return render_template('builder.html', step=1, proposal=dict(proposal), preselected_ids=[], map_workspaces=[])

@app.route('/proposals/<int:pid>/update', methods=['POST'])
def proposal_update(pid):
    data = request.json
    allowed_fields = ['title','template','client_name','client_company','client_email',
                      'client_location','team_size','space_type','area_required','budget',
                      'duration','selected_centres','manual_centres','status']
    sets = ', '.join(f'{f}=?' for f in allowed_fields if f in data)
    vals = [data[f] for f in allowed_fields if f in data] + [pid]
    if sets:
        with get_db() as conn:
            conn.execute(f'UPDATE proposals SET {sets} WHERE id=?', vals)
    return jsonify({'ok': True})

@app.route('/proposals/<int:pid>/delete', methods=['POST','DELETE'])
def proposal_delete(pid):
    with get_db() as conn:
        row = conn.execute('SELECT pptx_filename, pdf_filename FROM proposals WHERE id=?', (pid,)).fetchone()
        if row:
            for fname in [row['pptx_filename'], row['pdf_filename']]:
                if fname:
                    fpath = os.path.join(PROPOSAL_FILES, fname)
                    if os.path.exists(fpath): os.remove(fpath)
        conn.execute('DELETE FROM proposals WHERE id=?', (pid,))
    return jsonify({'ok': True})

@app.route('/proposals/<int:pid>/generate', methods=['POST'])
def proposal_generate(pid):
    with get_db() as conn:
        proposal = conn.execute('SELECT * FROM proposals WHERE id=?', (pid,)).fetchone()
        if not proposal:
            return jsonify({'error': 'Not found'}), 404
        p = dict(proposal)

    selected_ids = json.loads(p.get('selected_centres') or '[]')
    raw_manual = json.loads(p.get('manual_centres') or '[]')

    db_centres = []
    if selected_ids:
        with get_db() as conn:
            for cid in selected_ids:
                c = conn.execute('SELECT * FROM centres WHERE id=?', (cid,)).fetchone()
                if c:
                    imgs = conn.execute(
                        'SELECT filename FROM centre_images WHERE centre_id=? ORDER BY is_primary DESC, sort_order LIMIT 4',
                        (cid,)).fetchall()
                    centre_data = dict(c)
                    centre_data['images'] = [
                        r['filename'] if r['filename'].startswith('http')
                        else os.path.join(CENTRE_IMAGES, str(cid), r['filename'])
                        for r in imgs
                    ]
                    # Flag for coworking proposals: use hotdesk daily price in slides
                    if (p.get('space_type') or '').lower() == 'coworking':
                        centre_data['use_hotdesk'] = True
                    db_centres.append(centre_data)

    # Convert base64 images in manual centres to temp files; pass URL images directly
    tmp_files = []
    manual_centres = []
    tmp_dir = os.path.join(UPLOADS, 'tmp_render')
    os.makedirs(tmp_dir, exist_ok=True)
    for mc in raw_manual:
        b64_list = mc.get('images_b64') or []
        img_paths = []
        # URL images (from Hubble CDN via map)
        for url in (mc.get('image_urls') or mc.get('images') or [])[:4]:
            if isinstance(url, str) and url.startswith('http'):
                img_paths.append(url)
        for b64 in b64_list[:max(0, 4 - len(img_paths))]:
            if b64 and b64.startswith('data:image'):
                try:
                    _, data = b64.split(',', 1)
                    tmp_path = os.path.join(tmp_dir, f'tmp_{os.urandom(6).hex()}.jpg')
                    Image.open(io.BytesIO(base64.b64decode(data))).convert('RGB').save(tmp_path)
                    img_paths.append(tmp_path)
                    tmp_files.append(tmp_path)
                except:
                    pass
        mc_copy = {k: v for k, v in mc.items() if k != 'images_b64'}
        mc_copy['images'] = img_paths
        manual_centres.append(mc_copy)

    template = p.get('template','london')
    try:
        if template == 'india':
            slides = build_india_slides(p, db_centres, manual_centres)
        elif template == 'bold':
            slides = build_bold_slides(p, db_centres, manual_centres)
        else:
            slides = build_london_slides(p, db_centres, manual_centres)
        base_name = f'proposal_{pid}_{os.urandom(4).hex()}'
        pptx_path = render_pptx(slides, os.path.join(PROPOSAL_FILES, base_name + '.pptx'))
        pdf_path  = render_pdf(slides,  os.path.join(PROPOSAL_FILES, base_name + '.pdf'))
    except Exception as e:
        import traceback; traceback.print_exc()
        return jsonify({'error': str(e)}), 500
    finally:
        for f in tmp_files:
            try: os.remove(f)
            except: pass

    pptx_name = os.path.basename(pptx_path)
    pdf_name  = os.path.basename(pdf_path)
    with get_db() as conn:
        conn.execute('UPDATE proposals SET pptx_filename=?, pdf_filename=?, status=? WHERE id=?',
                     (pptx_name, pdf_name, 'generated', pid))
    return jsonify({'ok': True, 'pptx': pptx_name, 'pdf': pdf_name})

@app.route('/api/detect-template', methods=['POST'])
def api_detect_template():
    """Analyse an uploaded PPTX template and return feature metadata."""
    f = request.files.get('file')
    if not f:
        return jsonify({'error': 'No file'}), 400
    tmp = os.path.join(TEMPLATE_FILES, 'detect_tmp.pptx')
    f.save(tmp)
    features = detect_template_features(tmp)
    try:
        os.remove(tmp)
    except Exception:
        pass
    return jsonify(features)


@app.route('/proposals/<int:pid>/download')
@app.route('/proposals/<int:pid>/download/<fmt>')
def proposal_download(pid, fmt='pptx'):
    with get_db() as conn:
        row = conn.execute('SELECT pptx_filename, pdf_filename, title FROM proposals WHERE id=?', (pid,)).fetchone()
    if not row:
        return 'Not found', 404
    fname = row['pdf_filename'] if fmt == 'pdf' else row['pptx_filename']
    if not fname:
        return 'Not generated yet', 404
    fpath = os.path.join(PROPOSAL_FILES, fname)
    if not os.path.exists(fpath):
        return 'File not found', 404
    safe_title = (row['title'] or 'proposal').replace(' ','_')
    ext = 'pdf' if fmt == 'pdf' else 'pptx'
    return send_file(fpath, as_attachment=True, download_name=f'{safe_title}.{ext}')

# ── Slide-spec generation + dual rendering ───────────────────────────────────
# Delegate to generation.py which contains the full implementation.

from generation import (
    build_london_slides,
    build_india_slides,
    build_bold_slides,
    detect_template_features,
    render_pptx,
    render_pdf,
    get_logo_png,
)


# ── Routes: Share Links & Tracking ──────────────────────────────────────────

_RAILWAY_URL = 'https://web-production-df90c.up.railway.app'
BASE_URL = os.environ.get('BASE_URL', '')  # resolved dynamically per request below

def _self_warmup():
    import urllib.request as _ur
    while True:
        time.sleep(200)
        try:
            _ur.urlopen(f'{_RAILWAY_URL}/health', timeout=10)
        except Exception:
            pass

threading.Thread(target=_self_warmup, daemon=True).start()

def _base_url():
    """Use request host URL so links always work wherever Flask is running."""
    try:
        return request.host_url.rstrip('/')
    except Exception:
        return BASE_URL or _RAILWAY_URL

@app.route('/api/share-link/create', methods=['POST'])
def share_link_create():
    data = request.json or {}
    centre_ids = [str(c) for c in data.get('centre_ids', [])]
    label = data.get('label', '')
    client_email = data.get('client_email', '') or None
    client_phone = data.get('client_phone', '') or None
    if not centre_ids:
        return jsonify({'error': 'centre_ids required'}), 400
    centre_names = data.get('centre_names', [])
    sorted_ids = sorted(centre_ids)

    with get_db() as conn:
        # Resolve names from DB if not provided, and flag any id that doesn't match
        # a centre here at all — e.g. a space that exists on the map but was never
        # added to this tool's database. Without this check the link still gets
        # created, but the client-facing page silently renders with that space
        # missing (or empty entirely), which looks like "the proposal isn't there."
        names_provided = bool(centre_names)
        unmatched = []
        for idx, cid in enumerate(centre_ids):
            row = conn.execute(
                'SELECT name FROM centres WHERE hubble_id=? OR id=? LIMIT 1',
                (cid, cid)
            ).fetchone()
            if row:
                if not names_provided:
                    centre_names.append(row['name'])
            else:
                fallback_name = centre_names[idx] if names_provided and idx < len(centre_names) else cid
                unmatched.append(fallback_name)
                if not names_provided:
                    centre_names.append(fallback_name)

        # Reuse existing link with same space set (avoids dashboard duplicates)
        canonical_key = '|'.join(sorted_ids)
        token = None
        row = conn.execute(
            'SELECT token FROM share_links WHERE canonical_ids=? ORDER BY created_at DESC LIMIT 1',
            (canonical_key,)
        ).fetchone()
        if row:
            token = row['token']
        else:
            for old_row in conn.execute('SELECT token, centre_ids FROM share_links WHERE canonical_ids IS NULL ORDER BY created_at DESC LIMIT 200').fetchall():
                try:
                    if sorted(json.loads(old_row['centre_ids'])) == sorted_ids:
                        token = old_row['token']
                        conn.execute('UPDATE share_links SET canonical_ids=? WHERE token=?', (canonical_key, token))
                        break
                except Exception:
                    continue

        if not token:
            token = secrets.token_urlsafe(8)
            conn.execute(
                'INSERT INTO share_links (token, label, centre_ids, centre_names, canonical_ids, client_email, client_phone) VALUES (?,?,?,?,?,?,?)',
                (token, label, json.dumps(centre_ids), json.dumps(centre_names), canonical_key, client_email, client_phone)
            )
        elif label:
            conn.execute('UPDATE share_links SET label=? WHERE token=?', (label, token))

    return jsonify({
        'token': token,
        'url': f'{_base_url()}/compare/{token}',
        'label': label,
        'centre_names_list': centre_names,
        'client_email': client_email,
        'client_phone': client_phone,
        'unmatched': unmatched,
    })


_BAD_BRANDS = {'the','a','an','our','new','old','my','one','at','of','in','by'}

_ip_country_cache = {}  # ip → 'IN' / 'GB' / etc., cached in-process

def _is_india_ip(ip):
    """Returns True if the IP geolocates to India. Cached per IP, non-blocking."""
    if not ip or ip in ('127.0.0.1', '::1') or ip.startswith('192.168.') or ip.startswith('10.'):
        return False
    if ip in _ip_country_cache:
        return _ip_country_cache[ip] == 'IN'
    try:
        r = requests.get(f'http://ip-api.com/json/{ip}?fields=countryCode', timeout=2)
        country = r.json().get('countryCode', '')
    except Exception:
        country = ''
    _ip_country_cache[ip] = country
    return country == 'IN'

def _load_centres_for_compare(conn, hubble_ids):
    """Fetch centre data + images for a list of hubble_ids (batched)."""
    if not hubble_ids:
        return []
    ph = ','.join('?' * len(hubble_ids))
    rows = conn.execute(f'SELECT * FROM centres WHERE hubble_id IN ({ph})', hubble_ids).fetchall()
    # Also try by DB id for any that didn't match by hubble_id
    found_hids = {r['hubble_id'] for r in rows}
    fallback_ids = [h for h in hubble_ids if h not in found_hids]
    if fallback_ids:
        ph2 = ','.join('?' * len(fallback_ids))
        rows = list(rows) + conn.execute(f'SELECT * FROM centres WHERE id IN ({ph2})', fallback_ids).fetchall()

    centre_map = {r['hubble_id']: dict(r) for r in rows}
    # Fetch all images in one query
    cids = [c['id'] for c in centre_map.values()]
    if cids:
        ph3 = ','.join('?' * len(cids))
        img_rows = conn.execute(
            f'SELECT centre_id, filename FROM centre_images WHERE centre_id IN ({ph3}) ORDER BY is_primary DESC, sort_order',
            cids
        ).fetchall()
        imgs_by_centre = {}
        for ir in img_rows:
            imgs_by_centre.setdefault(ir['centre_id'], []).append(ir['filename'])

    centres = []
    for hid in hubble_ids:
        c = centre_map.get(hid)
        if not c:
            continue
        cid = c['id']
        raw_imgs = imgs_by_centre.get(cid, []) if cids else []
        c['image_urls'] = [
            fn if fn.startswith('http') else f'/centre-image/{cid}/{fn}'
            for fn in raw_imgs
        ]
        try:
            c['amenities_list'] = json.loads(c.get('amenities') or '[]')
        except Exception:
            c['amenities_list'] = []
        b = (c.get('brand') or '').strip()
        if b.lower() in _BAD_BRANDS or (b and b.isdigit()):
            c['brand'] = ''
        centres.append(c)
    return centres


@app.route('/compare')
def compare_direct():
    """Entry point from map — ?ids=hubble_id1,hubble_id2&names=name1,name2[&client=ClientName].
    Creates or reuses a stable token then REDIRECTS to /compare/<token> so that
    the client's browser URL is always the canonical token URL — no new links on refresh."""
    import secrets as _sec
    ids_param    = request.args.get('ids', '').strip()
    names_param  = request.args.get('names', '').strip()
    client_param = request.args.get('client', '').strip()
    price_mode   = request.args.get('price_mode', '').strip()
    personalised_message = request.args.get('personalised_message', '').strip()
    recommended_ids_param = request.args.get('recommended_ids', '').strip()
    if not ids_param:
        return 'No space IDs provided', 400

    hubble_ids = [i.strip() for i in ids_param.split(',') if i.strip()]
    names      = [urllib.parse.unquote(n) for n in names_param.split(',') if n.strip()]
    client_name = urllib.parse.unquote(client_param) if client_param else ''
    sorted_ids = sorted(hubble_ids)

    with get_db() as conn:
        centre_names = names or [str(h) for h in hubble_ids]

        # Reuse an existing share_link with the same set of space IDs
        # Use a canonical key (sorted, pipe-separated) to match without scanning all rows
        canonical_key = '|'.join(sorted_ids)
        token = None
        existing_label = None
        row = conn.execute(
            'SELECT token, label FROM share_links WHERE canonical_ids=? ORDER BY created_at DESC LIMIT 1',
            (canonical_key,)
        ).fetchone()
        if row:
            token = row['token']
            existing_label = row['label']
        else:
            # Legacy: scan recent links (capped) for backward compat with rows without canonical_ids
            for old_row in conn.execute('SELECT token, label, centre_ids FROM share_links WHERE canonical_ids IS NULL ORDER BY created_at DESC LIMIT 200').fetchall():
                try:
                    if sorted(json.loads(old_row['centre_ids'])) == sorted_ids:
                        token = old_row['token']
                        existing_label = old_row['label']
                        # Backfill canonical_ids
                        conn.execute('UPDATE share_links SET canonical_ids=? WHERE token=?', (canonical_key, token))
                        break
                except Exception:
                    continue

        if not token:
            token = _sec.token_urlsafe(8)
            if client_name:
                label = client_name
            else:
                label = ', '.join(centre_names[:2]) + (' + more' if len(centre_names) > 2 else '')
            conn.execute(
                'INSERT INTO share_links (token, label, centre_ids, centre_names, canonical_ids, personalised_message, recommended_ids) VALUES (?,?,?,?,?,?,?)',
                (token, label, json.dumps(hubble_ids), json.dumps(centre_names), canonical_key,
                 personalised_message or None, recommended_ids_param or None)
            )
        else:
            updates = []
            params = []
            if client_name and (not existing_label or existing_label != client_name):
                updates.append('label=?'); params.append(client_name)
            if personalised_message:
                updates.append('personalised_message=?'); params.append(personalised_message)
            if recommended_ids_param:
                updates.append('recommended_ids=?'); params.append(recommended_ids_param)
            if updates:
                conn.execute(f'UPDATE share_links SET {", ".join(updates)} WHERE token=?', params + [token])

        # Load data in the same connection — avoids a second round-trip
        link = conn.execute('SELECT * FROM share_links WHERE token=?', (token,)).fetchone()
        if not link:
            return 'Link not found', 500
        link = dict(link)
        hubble_ids = json.loads(link['centre_ids'])
        centres = _load_centres_for_compare(conn, hubble_ids)

    raw_ip = (request.headers.get('X-Forwarded-For') or request.remote_addr or '').split(',')[0].strip()
    ip_hash = hashlib.md5(raw_ip.encode()).hexdigest()[:8]
    ua = request.headers.get('User-Agent', '')

    def _log_open():
        if _is_india_ip(raw_ip):
            return
        try:
            with get_db() as c:
                # OR IGNORE + the partial unique index (see init_db) atomically drops a
                # same-second duplicate — a plain check-then-insert isn't safe here since
                # two near-simultaneous requests can both pass the check before either commits.
                c.execute(
                    'INSERT OR IGNORE INTO link_events (token, event_type, ip_hash, user_agent) VALUES (?,?,?,?)',
                    (token, 'open', ip_hash, ua)
                )
        except Exception:
            pass
        _push_sse({'token': token, 'event_type': 'open', 'centre_name': ''})

    rec_ids = json.loads(link.get('recommended_ids') or '[]')
    if rec_ids:
        rec_str = [str(r) for r in rec_ids]
        centres = sorted(centres, key=lambda c: 0 if str(c.get('id','')) in rec_str or str(c.get('hubble_id','')) in rec_str else 1)
    threading.Thread(target=_log_open, daemon=True).start()
    return render_template('compare.html', link=link, centres=centres, token=token,
                           price_mode=price_mode,
                           personalised_message=link.get('personalised_message') or '',
                           recommended_ids=rec_ids, is_preview=False,
                           canonical_url=url_for('compare_page', token=token, _external=True))


@app.route('/compare/<token>')
def compare_page(token):
    with get_db() as conn:
        link = conn.execute('SELECT * FROM share_links WHERE token=?', (token,)).fetchone()
        if not link:
            return 'Link not found', 404
        link = dict(link)
        hubble_ids = json.loads(link['centre_ids'])
        centres = _load_centres_for_compare(conn, hubble_ids)

    raw_ip  = (request.headers.get('X-Forwarded-For') or request.remote_addr or '').split(',')[0].strip()
    ip_hash = hashlib.md5(raw_ip.encode()).hexdigest()[:8]
    ua      = request.headers.get('User-Agent', '')

    def _log_open():
        if _is_india_ip(raw_ip):
            return
        try:
            with get_db() as c:
                # OR IGNORE + the partial unique index (see init_db) atomically drops a
                # same-second duplicate — a plain check-then-insert isn't safe here since
                # two near-simultaneous requests can both pass the check before either commits.
                c.execute(
                    'INSERT OR IGNORE INTO link_events (token, event_type, ip_hash, user_agent) VALUES (?,?,?,?)',
                    (token, 'open', ip_hash, ua)
                )
        except Exception:
            pass
        # Push SSE outside DB context so it always fires even if DB write fails
        _push_sse({'token': token, 'event_type': 'open', 'centre_name': ''})

    rec_ids = json.loads(link.get('recommended_ids') or '[]')
    if rec_ids:
        rec_str = [str(r) for r in rec_ids]
        centres = sorted(centres, key=lambda c: 0 if str(c.get('id','')) in rec_str or str(c.get('hubble_id','')) in rec_str else 1)
    is_preview = request.args.get('preview') == '1'
    if not is_preview:
        threading.Thread(target=_log_open, daemon=True).start()
    price_mode = request.args.get('price_mode', '').strip()
    return render_template('compare.html', link=link, centres=centres, token=token, price_mode=price_mode,
                           personalised_message=link.get('personalised_message') or '',
                           recommended_ids=rec_ids, is_preview=is_preview)


@app.route('/health')
def health():
    return jsonify({'ok': True})


@app.route('/track', methods=['POST'])
def track_event():
    # force=True parses JSON regardless of Content-Type (sendBeacon may send text/plain)
    data = request.get_json(force=True, silent=True) or {}
    token = data.get('token', '')
    event_type = data.get('event_type', '')
    if not token or not event_type:
        return jsonify({'ok': False, 'error': 'token and event_type required'}), 400
    ip = (request.headers.get('X-Forwarded-For') or request.remote_addr or '').split(',')[0].strip()
    if _is_india_ip(ip):
        return jsonify({'ok': True})
    ip_hash = hashlib.md5(ip.encode()).hexdigest()[:8]
    ua = request.headers.get('User-Agent', '')
    payload = {
        'token': token, 'event_type': event_type,
        'centre_id': data.get('centre_id'), 'centre_name': data.get('centre_name'),
        'dwell_seconds': data.get('dwell_seconds'), 'ip_hash': ip_hash, 'ua': ua,
        'booking_date': data.get('booking_date'), 'booking_time': data.get('booking_time'),
    }
    def _do_track():
        try:
            with get_db() as conn:
                # OR IGNORE + the partial unique index (see init_db) atomically drops a
                # same-second duplicate 'dwell' beacon — doesn't affect other event types.
                conn.execute(
                    '''INSERT OR IGNORE INTO link_events
                       (token, event_type, centre_id, centre_name, dwell_seconds, ip_hash, user_agent, booking_date, booking_time)
                       VALUES (?,?,?,?,?,?,?,?,?)''',
                    (payload['token'], payload['event_type'], payload['centre_id'],
                     payload['centre_name'], payload['dwell_seconds'], payload['ip_hash'],
                     payload['ua'], payload['booking_date'], payload['booking_time'])
                )
        except Exception:
            pass
    threading.Thread(target=_do_track, daemon=True).start()
    return jsonify({'ok': True})


@app.route('/dashboard/stream')
def dashboard_stream():
    # SSE replaced by polling — return 410 so old browser tabs stop retrying
    return jsonify({'error': 'use polling'}), 410


@app.route('/dashboard')
def dashboard():
    with get_db() as conn:
        links = conn.execute('SELECT * FROM share_links ORDER BY created_at DESC').fetchall()
        links = [dict(l) for l in links]

        counts = conn.execute(
            f"""SELECT token, event_type, COUNT(*) as cnt
               FROM ({_GENUINE_EVENTS_SQL}) WHERE event_type IN ('open','click')
               GROUP BY token, event_type"""
        ).fetchall()
        top_spaces = conn.execute(
            f"""SELECT token, centre_name, COUNT(*) as cnt
               FROM ({_GENUINE_EVENTS_SQL}) WHERE event_type='click' AND centre_name IS NOT NULL AND centre_name != ''
               GROUP BY token, centre_name"""
        ).fetchall()
        # Distinct spaces clicked, not raw click events — re-clicking the same
        # space (e.g. Fora 3 times) counts once, clicking Fora then WeWork counts as 2.
        unique_clicks_rows = conn.execute(
            f"""SELECT token, COUNT(DISTINCT COALESCE(centre_id, centre_name)) as cnt
               FROM ({_GENUINE_EVENTS_SQL})
               WHERE event_type='click' AND (centre_id IS NOT NULL OR centre_name IS NOT NULL)
               GROUP BY token"""
        ).fetchall()

        opens_map, clicks_map = {}, {}
        for row in counts:
            if row['event_type'] == 'open': opens_map[row['token']] = row['cnt']
            else: clicks_map[row['token']] = row['cnt']
        unique_clicks_map = {row['token']: row['cnt'] for row in unique_clicks_rows}

        top_map = {}
        for row in top_spaces:
            t = row['token']
            if t not in top_map or row['cnt'] > top_map[t][1]:
                top_map[t] = (row['centre_name'], row['cnt'])

        bookings_rows = conn.execute(
            f"""SELECT token, centre_name, booking_date, booking_time
               FROM ({_GENUINE_EVENTS_SQL})
               WHERE event_type='booking_request' AND booking_date IS NOT NULL
                 AND booking_date != '' AND booking_date != 'tbd'
               ORDER BY created_at ASC"""
        ).fetchall()
        bookings_map = {}
        for row in bookings_rows:
            t = row['token']
            if t not in bookings_map:
                bookings_map[t] = []
            bookings_map[t].append({
                'name': row['centre_name'] or '',
                'date': row['booking_date'] or '',
                'time': row['booking_time'] or ''
            })

        # Minimal per-link data — panel detail loaded lazily via /api/link-detail/<token>
        detail_data = {}
        for lnk in links:
            token = lnk['token']
            lnk['opens'] = opens_map.get(token, 0)
            lnk['clicks'] = clicks_map.get(token, 0)
            lnk['unique_clicks'] = unique_clicks_map.get(token, 0)
            lnk['top_space'] = top_map[token][0] if token in top_map else None
            lnk['bookings'] = bookings_map.get(token, [])
            try:
                lnk['centre_names_list'] = json.loads(lnk.get('centre_names') or '[]')
            except Exception:
                lnk['centre_names_list'] = []
            detail_data[token] = {
                'stats': {'opens': lnk['opens'], 'clicks': lnk['clicks'], 'unique_clicks': lnk['unique_clicks']},
                'space_stats': None,
                'events': None,
                'link': {
                    'label': lnk.get('label', ''),
                    'token': token,
                    'client_email': lnk.get('client_email'),
                    'client_phone': lnk.get('client_phone'),
                    'centre_names_list': lnk['centre_names_list'],
                },
            }

    return render_template('dashboard.html', links=links, base_url=_base_url(),
                           detail_data_json=json.dumps(detail_data))


@app.route('/dashboard/stats/<token>')
def dashboard_stats(token):
    with get_db() as conn:
        opens = conn.execute(f"SELECT COUNT(*) FROM ({_GENUINE_EVENTS_SQL}) WHERE token=? AND event_type='open'", (token,)).fetchone()[0]
        clicks = conn.execute(f"SELECT COUNT(*) FROM ({_GENUINE_EVENTS_SQL}) WHERE token=? AND event_type='click'", (token,)).fetchone()[0]
        top = conn.execute(f"SELECT centre_name, COUNT(*) as cnt FROM ({_GENUINE_EVENTS_SQL}) WHERE token=? AND event_type='click' AND centre_name IS NOT NULL GROUP BY centre_name ORDER BY cnt DESC LIMIT 1", (token,)).fetchone()
    return jsonify({'opens': opens, 'clicks': clicks, 'top_space': top[0] if top else None})


@app.route('/api/share-link/<token>/delete', methods=['POST'])
def share_link_delete(token):
    with get_db() as conn:
        conn.execute('DELETE FROM share_links WHERE token=?', (token,))
        conn.execute('DELETE FROM link_events WHERE token=?', (token,))
    return jsonify({'ok': True})


@app.route('/api/share-link/<token>/mark-test', methods=['POST'])
def share_link_mark_test(token):
    """Flag/unflag a link as test data — excluded from analytics rollups but
    still visible (and deletable) in the dashboard list."""
    data = request.json or {}
    is_test = 1 if data.get('is_test') else 0
    with get_db() as conn:
        conn.execute('UPDATE share_links SET is_test=? WHERE token=?', (is_test, token))
    return jsonify({'ok': True, 'is_test': bool(is_test)})


@app.route('/api/poll-updates')
def api_poll_updates():
    """Lightweight stats poll — replaces SSE so gunicorn threads stay free."""
    with get_db() as conn:
        counts = conn.execute(
            f"""SELECT token, event_type, COUNT(*) as cnt
               FROM ({_GENUINE_EVENTS_SQL}) WHERE event_type IN ('open','click')
               GROUP BY token, event_type"""
        ).fetchall()
        top = conn.execute(
            f"""SELECT token, centre_name, COUNT(*) as cnt
               FROM ({_GENUINE_EVENTS_SQL}) WHERE event_type='click' AND centre_name IS NOT NULL AND centre_name != ''
               GROUP BY token, centre_name"""
        ).fetchall()
        unique_clicks = conn.execute(
            f"""SELECT token, COUNT(DISTINCT COALESCE(centre_id, centre_name)) as cnt
               FROM ({_GENUINE_EVENTS_SQL})
               WHERE event_type='click' AND (centre_id IS NOT NULL OR centre_name IS NOT NULL)
               GROUP BY token"""
        ).fetchall()
    stats = {}
    for r in counts:
        t = r['token']
        stats.setdefault(t, {'opens':0,'clicks':0,'unique_clicks':0,'top':None})
        if r['event_type'] == 'open': stats[t]['opens'] = r['cnt']
        else: stats[t]['clicks'] = r['cnt']
    for r in unique_clicks:
        t = r['token']
        stats.setdefault(t, {'opens':0,'clicks':0,'unique_clicks':0,'top':None})
        stats[t]['unique_clicks'] = r['cnt']
    top_map = {}
    for r in top:
        t = r['token']
        if t not in top_map or r['cnt'] > top_map[t][1]:
            top_map[t] = (r['centre_name'], r['cnt'])
    for t, v in top_map.items():
        if t in stats: stats[t]['top'] = v[0]
    return jsonify(stats)


@app.route('/api/link-events/<token>')
def api_link_events(token):
    """Lazy-loaded activity feed for the detail panel."""
    with get_db() as conn:
        rows = conn.execute(
            f"""SELECT event_type, centre_name, created_at FROM ({_GENUINE_EVENTS_SQL})
               WHERE token=? ORDER BY created_at DESC LIMIT 30""",
            (token,)
        ).fetchall()
    events = [dict(r) for r in rows
              if not (r['centre_name'] and r['centre_name'].isdigit())]
    return jsonify({'events': events})


@app.route('/api/visitor-detail')
def api_visitor_detail():
    token = request.args.get('token', '')
    ip_hash = request.args.get('ip_hash', '')
    if not token or not ip_hash:
        return jsonify({'error': 'token and ip_hash required'}), 400
    with get_db() as conn:
        detail = _get_visitor_detail(conn, token, ip_hash)
    if not detail:
        return jsonify({'error': 'not found'}), 404
    return jsonify(detail)


@app.route('/api/link-detail/<token>')
def api_link_detail(token):
    with get_db() as conn:
        link = conn.execute('SELECT * FROM share_links WHERE token=?', (token,)).fetchone()
        if not link:
            return jsonify({'error': 'not found'}), 404
        link = dict(link)
        try:
            link['centre_names_list'] = json.loads(link.get('centre_names') or '[]')
        except Exception:
            link['centre_names_list'] = []
        events = conn.execute(
            f'SELECT event_type, centre_name, centre_id, created_at, dwell_seconds, booking_date, booking_time FROM ({_GENUINE_EVENTS_SQL}) WHERE token=? ORDER BY created_at DESC LIMIT 200',
            (token,)
        ).fetchall()
        events = [dict(e) for e in events]
        space_stats = {}
        for ev in events:
            cn = ev.get('centre_name') or ev.get('centre_id') or ''
            if not cn:
                continue
            if cn not in space_stats:
                space_stats[cn] = {'clicks': 0, 'interested': 0, 'not_interested': 0, 'bookings': []}
            t = ev.get('event_type', '')
            if t == 'click':
                space_stats[cn]['clicks'] += 1
            elif t == 'interested':
                space_stats[cn]['interested'] += 1
            elif t == 'not_interested':
                space_stats[cn]['not_interested'] += 1
            elif t == 'booking_request' and ev.get('booking_date'):
                space_stats[cn]['bookings'].append(f"{ev['booking_date']} {ev.get('booking_time','')}")
        opens = sum(1 for e in events if e['event_type'] == 'open')
        clicks = sum(1 for e in events if e['event_type'] == 'click')
        unique_clicks = sum(1 for v in space_stats.values() if v['clicks'] > 0)
    return jsonify({
        'link': {k: link[k] for k in ('label','token','client_email','client_phone','created_at','centre_names_list') if k in link},
        'stats': {'opens': opens, 'clicks': clicks, 'unique_clicks': unique_clicks},
        'space_stats': [{'name': k, **v} for k, v in space_stats.items()],
        'events': events[:50],
    })


@app.route('/dashboard/<token>')
def dashboard_detail(token):
    with get_db() as conn:
        link = conn.execute('SELECT * FROM share_links WHERE token=?', (token,)).fetchone()
        if not link:
            return 'Link not found', 404
        link = dict(link)
        events = conn.execute(
            f'SELECT * FROM ({_GENUINE_EVENTS_SQL}) WHERE token=? ORDER BY created_at DESC',
            (token,)
        ).fetchall()
        events = [dict(e) for e in events]
        # Bar chart data: clicks per space
        click_rows = conn.execute(
            f"""SELECT centre_name, COUNT(*) as cnt FROM ({_GENUINE_EVENTS_SQL})
               WHERE token=? AND event_type='click' AND centre_name IS NOT NULL
               GROUP BY centre_name ORDER BY cnt DESC""",
            (token,)
        ).fetchall()
        chart_labels = [r['centre_name'] for r in click_rows]
        chart_data = [r['cnt'] for r in click_rows]
        try:
            link['centre_names_list'] = json.loads(link.get('centre_names') or '[]')
        except Exception:
            link['centre_names_list'] = []

        # Skip loading full centre profile/images — analytics page doesn't need them
        centres = []

        # Per-space stats: clicks, interest, bookings
        space_stats = {}
        for ev in events:
            cn = ev.get('centre_name') or ev.get('centre_id', '')
            if not cn:
                continue
            if cn not in space_stats:
                space_stats[cn] = {'clicks': 0, 'interested': 0, 'not_interested': 0, 'bookings': []}
            t = ev.get('event_type', '')
            if t == 'click':
                space_stats[cn]['clicks'] += 1
            elif t == 'interested':
                space_stats[cn]['interested'] += 1
            elif t == 'not_interested':
                space_stats[cn]['not_interested'] += 1
            elif t == 'booking_request' and ev.get('booking_date'):
                space_stats[cn]['bookings'].append(
                    f"{ev['booking_date']} {ev.get('booking_time','')}"
                )
        # Attach stats to centres
        for c in centres:
            stats = space_stats.get(c['name']) or space_stats.get(str(c.get('hubble_id',''))) or {}
            c['_clicks']       = stats.get('clicks', 0)
            c['_interested']   = stats.get('interested', 0)
            c['_not_interested'] = stats.get('not_interested', 0)
            c['_bookings']     = stats.get('bookings', [])

    return render_template('dashboard_detail.html', link=link, events=events,
                           chart_labels=chart_labels, chart_data=chart_data,
                           centres=centres, base_url=_base_url(), token=token)


_LONDON_TZ = ZoneInfo('Europe/London')

def _parse_utc(ts_str):
    """link_events.created_at is SQLite's CURRENT_TIMESTAMP — naive but actually UTC.
    Attach tzinfo explicitly so conversion to London time (and any other local
    time) accounts for BST/GMT correctly instead of being off by an hour half the year."""
    return datetime.strptime(ts_str, '%Y-%m-%d %H:%M:%S').replace(tzinfo=_dt_timezone.utc)

def _relative_time(dt_utc):
    secs = (datetime.now(_dt_timezone.utc) - dt_utc).total_seconds()
    if secs < 3600:
        return f'{max(1, int(secs // 60))} min ago'
    if secs < 86400:
        h = int(secs // 3600)
        return f'{h} hour{"s" if h != 1 else ""} ago'
    d = int(secs // 86400)
    return f'{d} day{"s" if d != 1 else ""} ago'

_IST_TZ = _dt_timezone(timedelta(hours=5, minutes=30))  # IST has no DST, unlike London

def _fmt_time_ampm(dt_local):
    return dt_local.strftime('%I:%M %p').lstrip('0')  # "09:15 AM" -> "9:15 AM"

def _fmt_date_part(dt_local):
    return f'{dt_local.strftime("%a")} {dt_local.day} {dt_local.strftime("%b")}'

def _fmt_dual_tz(dt_utc):
    """Every absolute clock-time shown to this India-based team needs to say
    which zone it's in — shows both London's current zone (BST or GMT,
    whichever actually applies on that date) and IST side by side, e.g.
    'Thu 20 Aug, 10:22 AM BST · 2:52 PM IST'."""
    london = dt_utc.astimezone(_LONDON_TZ)
    ist = dt_utc.astimezone(_IST_TZ)
    london_str = f'{_fmt_date_part(london)}, {_fmt_time_ampm(london)} {london.tzname()}'
    if ist.date() == london.date():
        ist_str = f'{_fmt_time_ampm(ist)} IST'
    else:
        ist_str = f'{_fmt_date_part(ist)}, {_fmt_time_ampm(ist)} IST'
    return f'{london_str} · {ist_str}'

# Links flagged is_test are excluded from every analytics rollup below —
# they stay visible/manageable in the dashboard list, just not counted.
_REAL_LINKS_SQL = "SELECT token FROM share_links WHERE is_test=0 OR is_test IS NULL"

# The creator always opens their own share link once to check it before sending
# it on — that first open (and anything clicked/dwelled during it) is the
# creator's own activity, not the client's, and would otherwise pollute every
# opens/clicks/dwell number below. Per-token, the cutoff is the SECOND 'open'
# event ever logged; everything from that point on is treated as genuine
# client activity. A link with fewer than 2 opens has no genuine data yet.
_SECOND_OPEN_CUTOFF_SQL = """
    SELECT token, created_at AS cutoff_at, id AS cutoff_id FROM (
        SELECT token, created_at, id,
               ROW_NUMBER() OVER (PARTITION BY token ORDER BY created_at, id) AS rn
        FROM link_events WHERE event_type='open'
    ) WHERE rn = 2
"""
_GENUINE_EVENTS_SQL = f"""
    SELECT le.* FROM link_events le
    JOIN ({_SECOND_OPEN_CUTOFF_SQL}) co ON co.token = le.token
    WHERE (le.created_at, le.id) >= (co.cutoff_at, co.cutoff_id)
"""

# Velocity-based hot-lead thresholds: (hours since their first open, opens
# needed within that window). Faster re-opening is a stronger buying signal
# than the same open count spread over a longer time, so these are checked
# fastest-window-first and the first one met wins.
HOT_LEAD_TIERS = [(24, 2), (48, 3), (72, 4)]

def _get_repeat_openers(conn, limit=10):
    """Clients who've opened the same link 2+ times and haven't booked yet —
    still comparing, not gone cold. 'hot' (met one of HOT_LEAD_TIERS) always
    sorts first regardless of recency — that velocity is the signal itself,
    not just a recent one. Within a tier, most recent activity first."""
    rows = conn.execute(f"""
        SELECT token, ip_hash, COUNT(*) as opens, MAX(created_at) as last_open
        FROM ({_GENUINE_EVENTS_SQL}) WHERE event_type='open' AND ip_hash IS NOT NULL
        AND token IN ({_REAL_LINKS_SQL})
        GROUP BY token, ip_hash HAVING COUNT(*) >= 2
        ORDER BY last_open DESC LIMIT ?
    """, (limit * 3,)).fetchall()  # over-fetch — some get dropped below if already booked

    results = []
    for r in rows:
        token, ip_hash = r['token'], r['ip_hash']
        has_booking = conn.execute(
            f"SELECT 1 FROM ({_GENUINE_EVENTS_SQL}) WHERE token=? AND ip_hash=? AND event_type='booking_request' LIMIT 1",
            (token, ip_hash)
        ).fetchone()
        if has_booking:
            continue  # already moved forward — not a "still deciding" case
        link = conn.execute('SELECT label FROM share_links WHERE token=?', (token,)).fetchone()
        # Distinct spaces clicked, not raw click events — re-clicking the same
        # space (e.g. to look again) must not inflate the count.
        clicks = conn.execute(
            f"""SELECT COUNT(DISTINCT COALESCE(centre_id, centre_name))
               FROM ({_GENUINE_EVENTS_SQL}) WHERE token=? AND ip_hash=? AND event_type='click'""",
            (token, ip_hash)
        ).fetchone()[0]

        # Velocity check needs every open timestamp for this visitor, not just
        # the count/last — "2 opens in the first 24h" depends on when their
        # FIRST open was, which COUNT(*)/MAX(created_at) alone can't answer.
        open_rows = conn.execute(
            f"""SELECT created_at FROM ({_GENUINE_EVENTS_SQL})
               WHERE token=? AND ip_hash=? AND event_type='open' ORDER BY created_at ASC""",
            (token, ip_hash)
        ).fetchall()
        open_dts = [_parse_utc(row['created_at']) for row in open_rows]
        first_open_dt = open_dts[0]
        last_open_dt = open_dts[-1]
        hours_since = (datetime.now(_dt_timezone.utc) - last_open_dt).total_seconds() / 3600

        action, hot_reason, hot_tier_hours = 'watch', None, None
        for hours, min_opens in HOT_LEAD_TIERS:
            cutoff = first_open_dt + timedelta(hours=hours)
            count_within = sum(1 for dt in open_dts if dt <= cutoff)
            if count_within >= min_opens:
                action = 'hot'
                hot_reason = f'{count_within} opens within {hours}h'
                hot_tier_hours = hours
                break
        if action != 'hot':
            action = 'call' if hours_since <= 24 else 'watch'

        results.append({
            'label': (link['label'] if link else None) or 'Unlabelled link',
            'token': token,
            'ip_hash': ip_hash,
            'opens': r['opens'],
            'last_open_dt': last_open_dt,
            'last_open_rel': _relative_time(last_open_dt),
            'clicks': clicks,
            'action': action,
            'hot_reason': hot_reason,
            'hot_tier_hours': hot_tier_hours,
        })

    tier_rank = {'hot': 0, 'call': 1, 'watch': 2}
    results.sort(key=lambda o: (tier_rank[o['action']], -o['last_open_dt'].timestamp()))
    for o in results:
        del o['last_open_dt']  # was only needed for sorting, not JSON/template-safe
    return results[:limit]

def _get_visitor_detail(conn, token, ip_hash):
    """Full activity for one visitor on one link: last open, distinct spaces
    clicked (not raw click count), total time spent (deduped, outliers capped
    the same way the analytics-page average is), and interest/booking signals."""
    link = conn.execute('SELECT label FROM share_links WHERE token=?', (token,)).fetchone()
    events = conn.execute(
        f"""SELECT event_type, centre_id, centre_name, dwell_seconds, created_at
           FROM ({_GENUINE_EVENTS_SQL}) WHERE token=? AND ip_hash=? ORDER BY created_at ASC""",
        (token, ip_hash)
    ).fetchall()
    if not events:
        return None

    opens = [e for e in events if e['event_type'] == 'open']
    clicked = {}     # key -> centre_name, de-duplicated
    interested = {}
    not_interested = {}
    total_dwell = 0
    bookings = []

    for e in events:
        key = e['centre_id'] or e['centre_name']
        if e['event_type'] == 'click' and key:
            clicked[key] = e['centre_name'] or key
        elif e['event_type'] == 'interested' and key:
            interested[key] = e['centre_name'] or key
        elif e['event_type'] == 'not_interested' and key:
            not_interested[key] = e['centre_name'] or key
        elif e['event_type'] == 'dwell' and e['dwell_seconds']:
            total_dwell += min(e['dwell_seconds'], 1800)  # same 30-min cap as the analytics average
        elif e['event_type'] == 'booking_request':
            bookings.append(e['centre_name'])

    last_open_dt = _parse_utc(opens[-1]['created_at']) if opens else _parse_utc(events[-1]['created_at'])
    first_open_dt = _parse_utc(opens[0]['created_at']) if opens else _parse_utc(events[0]['created_at'])

    return {
        'label': (link['label'] if link else None) or 'Unlabelled link',
        'opens': len(opens),
        'first_open_rel': _relative_time(first_open_dt),
        'last_open_rel': _relative_time(last_open_dt),
        'total_dwell_seconds': total_dwell,
        'spaces_clicked': list(clicked.values()),
        'spaces_interested': list(interested.values()),
        'spaces_not_interested': list(not_interested.values()),
        'bookings': bookings,
    }

def _get_peak_open_hours(conn):
    """Hour-of-day distribution of link opens, in London local time (not the
    raw UTC storage) — otherwise the 'peak hour' would be wrong by 0-1h
    depending on the time of year (BST vs GMT).

    Counts each link's genuine first open only (the SECOND open overall —
    the first is always the creator checking their own link before sending
    it) — a client re-opening the same link later shouldn't get counted
    again here, since this chart answers "when do clients first look at
    what I sent," not "when is anyone ever looking.\""""
    rows = conn.execute(f"""
        SELECT cutoff_at as created_at FROM ({_SECOND_OPEN_CUTOFF_SQL}) co
        WHERE co.token IN ({_REAL_LINKS_SQL})
    """).fetchall()
    hour_counts = [0] * 24       # UK local hour (BST or GMT depending on date)
    hour_counts_ist = [0] * 24   # same events, bucketed by IST hour instead
    for r in rows:
        try:
            dt_utc = _parse_utc(r['created_at'])
        except ValueError:
            continue
        hour_counts[dt_utc.astimezone(_LONDON_TZ).hour] += 1
        hour_counts_ist[dt_utc.astimezone(_IST_TZ).hour] += 1

    def _fmt_hour(h):
        suffix = 'am' if h % 24 < 12 else 'pm'
        hh = h % 12 or 12
        return f'{hh}{suffix}'

    def _peak_window(counts):
        if sum(counts) < 5:  # not enough data to draw a conclusion below this
            return None
        threshold = max(counts) * 0.7
        peak_hours = [h for h, c in enumerate(counts) if c >= threshold and c > 0]
        if not peak_hours:
            return None
        return f'{_fmt_hour(min(peak_hours))}–{_fmt_hour(max(peak_hours) + 1)}'

    peak_window = _peak_window(hour_counts)
    peak_window_ist = _peak_window(hour_counts_ist)
    return hour_counts, peak_window, peak_window_ist


def _fmt_short_date(d):
    """'Aug 4' — cross-platform equivalent of %-d (Windows has no such code)."""
    return f'{d.strftime("%b")} {d.day}'

def _get_weekly_report(conn):
    """Rolls engagement up into Monday–Sunday weeks (London time), covering
    every week from the first tracked activity through the current week —
    no gaps, so 'Week 3' always means the same thing across visits even if
    a week had zero activity. Week 1 is the earliest week on record."""
    opens = conn.execute(f"""
        SELECT created_at, token, ip_hash FROM ({_GENUINE_EVENTS_SQL})
        WHERE event_type='open' AND token IN ({_REAL_LINKS_SQL})
    """).fetchall()
    clicks = conn.execute(f"""
        SELECT created_at, centre_id, centre_name FROM ({_GENUINE_EVENTS_SQL})
        WHERE event_type='click' AND (centre_id IS NOT NULL OR centre_name IS NOT NULL)
        AND token IN ({_REAL_LINKS_SQL})
    """).fetchall()
    links = conn.execute("""
        SELECT token, created_at FROM share_links WHERE is_test=0 OR is_test IS NULL
    """).fetchall()

    def week_start(dt_utc):
        local = dt_utc.astimezone(_LONDON_TZ)
        return local.date() - timedelta(days=local.weekday())

    buckets = {}
    def bucket(d):
        return buckets.setdefault(d, {
            'total_opens': 0, 'client_opens': set(),
            'total_clicks': 0, 'spaces_clicked': set(),
            'proposals_sent': 0,
        })

    all_weeks_seen = []
    for r in opens:
        try:
            d = week_start(_parse_utc(r['created_at']))
        except ValueError:
            continue
        b = bucket(d)
        b['total_opens'] += 1
        b['client_opens'].add((r['token'], r['ip_hash']))
        all_weeks_seen.append(d)
    for r in clicks:
        try:
            d = week_start(_parse_utc(r['created_at']))
        except ValueError:
            continue
        b = bucket(d)
        b['total_clicks'] += 1
        b['spaces_clicked'].add(r['centre_id'] or r['centre_name'])
        all_weeks_seen.append(d)
    for r in links:
        try:
            d = week_start(_parse_utc(r['created_at']))
        except ValueError:
            continue
        bucket(d)['proposals_sent'] += 1
        all_weeks_seen.append(d)

    if not all_weeks_seen:
        return []

    first_week = min(all_weeks_seen)
    current_week = week_start(datetime.now(_dt_timezone.utc))
    last_week = max(first_week, current_week)

    weeks = []
    d, idx = first_week, 1
    empty = {'total_opens': 0, 'client_opens': set(), 'total_clicks': 0,
             'spaces_clicked': set(), 'proposals_sent': 0}
    while d <= last_week:
        b = buckets.get(d, empty)
        week_end = d + timedelta(days=6)
        weeks.append({
            'week_num': idx,
            'start': d.isoformat(),
            'label': f'{_fmt_short_date(d)} – {_fmt_short_date(week_end)}, {week_end.year}',
            'total_opens': b['total_opens'],
            'unique_client_opens': len(b['client_opens']),
            'total_clicks': b['total_clicks'],
            'unique_spaces_clicked': len(b['spaces_clicked']),
            'proposals_sent': b['proposals_sent'],
        })
        d += timedelta(days=7)
        idx += 1
    return weeks


@app.route('/dashboard/weekly-report')
def weekly_report():
    with get_db() as conn:
        weeks = _get_weekly_report(conn)
    return render_template('weekly_report.html', weeks=weeks)


def _parse_week_start_arg():
    """Returns (date, error_response). error_response is None on success."""
    start_str = request.args.get('start', '')
    try:
        return datetime.strptime(start_str, '%Y-%m-%d').date(), None
    except ValueError:
        return None, (jsonify({'error': 'invalid or missing start date'}), 400)

def _week_open_rows(conn, week_start):
    """Every genuine open whose London-local week matches week_start, as
    (token, ip_hash, label, dt_utc) — the shared basis for both the raw
    Total Opens drill-down and the per-client Unique Client Opens one."""
    rows = conn.execute(f"""
        SELECT le.created_at, le.token, le.ip_hash, sl.label
        FROM ({_GENUINE_EVENTS_SQL}) le
        JOIN share_links sl ON sl.token = le.token
        WHERE le.event_type='open' AND le.token IN ({_REAL_LINKS_SQL})
    """).fetchall()
    out = []
    for r in rows:
        try:
            dt_utc = _parse_utc(r['created_at'])
        except ValueError:
            continue
        local = dt_utc.astimezone(_LONDON_TZ)
        if local.date() - timedelta(days=local.weekday()) != week_start:
            continue
        out.append({'token': r['token'], 'ip_hash': r['ip_hash'],
                     'label': r['label'] or 'Unlabelled link', 'dt_utc': dt_utc})
    return out

@app.route('/api/weekly-report/opens')
def api_weekly_report_opens():
    """Every genuine open in one Monday-Sunday week, most recent first —
    the drill-down behind clicking 'Total Opens' on the weekly report."""
    week_start, err = _parse_week_start_arg()
    if err:
        return err
    with get_db() as conn:
        rows = _week_open_rows(conn, week_start)
    events = sorted(rows, key=lambda r: r['dt_utc'], reverse=True)
    return jsonify({'events': [
        {'when': _fmt_dual_tz(r['dt_utc']), 'client': r['label'], 'token': r['token']}
        for r in events
    ]})

@app.route('/api/weekly-report/unique-client-opens')
def api_weekly_report_unique_client_opens():
    """One row per distinct (link, visitor) that opened during the week,
    each showing how many times they opened it and their most recent open —
    the drill-down behind clicking 'Unique Client Opens'. This is the same
    set _get_weekly_report counts via len(client_opens) for that week."""
    week_start, err = _parse_week_start_arg()
    if err:
        return err
    with get_db() as conn:
        rows = _week_open_rows(conn, week_start)

    clients = {}
    for r in rows:
        key = (r['token'], r['ip_hash'])
        c = clients.setdefault(key, {'label': r['label'], 'token': r['token'], 'opens': 0, 'last_dt': r['dt_utc']})
        c['opens'] += 1
        if r['dt_utc'] > c['last_dt']:
            c['last_dt'] = r['dt_utc']

    ordered = sorted(clients.values(), key=lambda c: c['last_dt'], reverse=True)
    return jsonify({'events': [
        {'client': c['label'], 'token': c['token'], 'opens': c['opens'], 'when': _fmt_dual_tz(c['last_dt'])}
        for c in ordered
    ]})


@app.route('/dashboard/analytics')
def dashboard_analytics():
    from collections import Counter
    with get_db() as conn:
        # Top centres by click count
        top_centres = conn.execute(f"""
            SELECT centre_name, COUNT(*) as clicks
            FROM ({_GENUINE_EVENTS_SQL}) WHERE event_type='click' AND centre_name IS NOT NULL
            AND token IN ({_REAL_LINKS_SQL})
            GROUP BY centre_name ORDER BY clicks DESC LIMIT 10
        """).fetchall()

        # Top brands — extract brand prefix (before first ' - ' or ' – ') via SQL, aggregate in Python
        brand_raw_rows = conn.execute(f"""
            SELECT TRIM(SUBSTR(centre_name, 1,
                CASE WHEN INSTR(centre_name,' - ')>0 THEN INSTR(centre_name,' - ')-1
                     WHEN INSTR(centre_name,' – ')>0 THEN INSTR(centre_name,' – ')-1
                     ELSE LENGTH(centre_name) END
            )) as brand, COUNT(*) as cnt
            FROM ({_GENUINE_EVENTS_SQL})
            WHERE event_type='click' AND centre_name IS NOT NULL
            AND token IN ({_REAL_LINKS_SQL})
            GROUP BY brand
        """).fetchall()

        # Avg dwell time (in seconds)
        # Cap at 30 min — a browser tab left open in the background racks up a huge
        # dwell_seconds value that isn't real reading time, and one such session can
        # drag the whole average up to something implausible (we saw 126 min average
        # driven by a single 8.5-hour "open tab" outlier).
        avg_dwell = conn.execute(f"""
            SELECT AVG(dwell_seconds) FROM ({_GENUINE_EVENTS_SQL})
            WHERE event_type='dwell' AND dwell_seconds IS NOT NULL
            AND dwell_seconds > 5 AND dwell_seconds <= 1800
            AND token IN ({_REAL_LINKS_SQL})
        """).fetchone()[0]

        # Total opens, clicks, unique links — single query
        totals = conn.execute(f"""
            SELECT
                SUM(CASE WHEN event_type='open' THEN 1 ELSE 0 END) as opens,
                SUM(CASE WHEN event_type='click' THEN 1 ELSE 0 END) as clicks,
                (SELECT COUNT(DISTINCT token) FROM ({_SECOND_OPEN_CUTOFF_SQL}) co
                 WHERE co.token IN ({_REAL_LINKS_SQL})) as unique_links
            FROM ({_GENUINE_EVENTS_SQL})
            WHERE token IN ({_REAL_LINKS_SQL})
        """).fetchone()
        total_opens = totals[0] or 0
        total_clicks = totals[1] or 0
        unique_links = totals[2] or 0

        # Booking time preferences
        time_prefs = conn.execute(f"""
            SELECT booking_time, COUNT(*) as cnt FROM ({_GENUINE_EVENTS_SQL})
            WHERE event_type='booking_request' AND booking_time IS NOT NULL
            AND token IN ({_REAL_LINKS_SQL})
            GROUP BY booking_time ORDER BY cnt DESC LIMIT 8
        """).fetchall()

        # Avg spaces per link and avg clicks per link
        avg_spaces = conn.execute("""
            SELECT AVG(json_array_length(centre_ids)) FROM share_links
            WHERE centre_ids IS NOT NULL AND (is_test=0 OR is_test IS NULL)
        """).fetchone()[0]
        avg_clicks_per_link = conn.execute(f"""
            SELECT AVG(c) FROM (
                SELECT COUNT(*) as c FROM ({_GENUINE_EVENTS_SQL})
                WHERE event_type='click' AND token IN ({_REAL_LINKS_SQL})
                GROUP BY token
            )
        """).fetchone()[0]

        repeat_openers = _get_repeat_openers(conn)
        hour_counts, peak_window, peak_window_ist = _get_peak_open_hours(conn)

    # Group hot leads by which velocity tier they tripped, fastest first, so
    # "2+ in 24h" / "3+ in 48h" / "4+ in 72h" can each be shown as their own table.
    hot_by_tier = [
        {
            'hours': hours,
            'min_opens': min_opens,
            'leads': [o for o in repeat_openers if o.get('hot_tier_hours') == hours],
        }
        for hours, min_opens in HOT_LEAD_TIERS
    ]

    # Normalise brand names
    brand_counts = Counter()
    for brand_raw, cnt in brand_raw_rows:
        b = (brand_raw or '').strip()
        bl = b.lower()
        if 'wework' in bl: b = 'WeWork'
        elif 'fora' in bl: b = 'Fora'
        elif 'regus' in bl: b = 'Regus'
        elif 'spaces' in bl and 'co' not in bl: b = 'Spaces'
        elif 'iwg' in bl: b = 'IWG'
        elif 'workspace' in bl: b = 'Workspace'
        elif 'labs' in bl: b = 'LABS'
        elif 'x+why' in bl or 'x why' in bl or 'xwhy' in bl: b = 'x+why'
        elif 'runway' in bl: b = 'Runway East'
        elif 'uncommon' in bl: b = 'Uncommon'
        elif 'the brew' in bl: b = 'The Brew'
        elif 'huckletree' in bl: b = 'Huckletree'
        elif 'trampery' in bl: b = 'The Trampery'
        if b:
            brand_counts[b] += cnt
    top_brands = brand_counts.most_common(8)

    return render_template('analytics.html',
        top_centres=[dict(r) for r in top_centres],
        top_brands=top_brands,
        avg_dwell=round(avg_dwell or 0),
        total_opens=total_opens,
        total_clicks=total_clicks,
        unique_links=unique_links,
        time_prefs=[dict(r) for r in time_prefs],
        avg_spaces=round(avg_spaces or 0, 1),
        avg_clicks_per_link=round(avg_clicks_per_link or 0, 1),
        repeat_openers=repeat_openers,
        hot_by_tier=hot_by_tier,
        hour_counts=hour_counts,
        peak_window=peak_window,
        peak_window_ist=peak_window_ist,
        HOT_LEAD_TIERS=HOT_LEAD_TIERS,
    )


# ── Backup / Restore (share links) ──────────────────────────────────────────
@app.route('/admin/backup')
def admin_backup_page():
    return render_template('backup.html')

def _build_backup_payload(conn):
    conn.row_factory = sqlite3.Row
    return {
        'centres': [dict(r) for r in conn.execute('SELECT * FROM centres ORDER BY id').fetchall()],
        'centre_images': [dict(r) for r in conn.execute('SELECT * FROM centre_images ORDER BY id').fetchall()],
        'proposals': [dict(r) for r in conn.execute('SELECT * FROM proposals ORDER BY id').fetchall()],
        'share_links': [dict(r) for r in conn.execute('SELECT * FROM share_links ORDER BY created_at').fetchall()],
        'link_events': [dict(r) for r in conn.execute('SELECT * FROM link_events ORDER BY created_at').fetchall()],
    }

@app.route('/admin/backup/export')
def admin_backup_export():
    with get_db() as conn:
        payload = _build_backup_payload(conn)
    body = json.dumps(payload, default=str, indent=2)
    return send_file(
        io.BytesIO(body.encode()),
        mimetype='application/json',
        as_attachment=True,
        download_name=f'proposal_tool_backup_{datetime.now(_dt_timezone.utc).strftime("%Y%m%d")}.json'
    )

_AUTO_BACKUP_DIR = os.path.join(_DATA_DIR, 'backups')
_AUTO_BACKUP_KEEP_DAYS = 14
_AUTO_BACKUP_INTERVAL_SECONDS = 24 * 60 * 60

def _run_scheduled_backup():
    os.makedirs(_AUTO_BACKUP_DIR, exist_ok=True)
    with get_db() as conn:
        payload = _build_backup_payload(conn)
    stamp = datetime.now(_dt_timezone.utc).strftime('%Y-%m-%d')
    path = os.path.join(_AUTO_BACKUP_DIR, f'auto_backup_{stamp}.json')
    with open(path, 'w', encoding='utf-8') as fh:
        json.dump(payload, fh, default=str)
    cutoff = time.time() - _AUTO_BACKUP_KEEP_DAYS * 86400
    for fname in os.listdir(_AUTO_BACKUP_DIR):
        fpath = os.path.join(_AUTO_BACKUP_DIR, fname)
        if fname.startswith('auto_backup_') and os.path.isfile(fpath) and os.path.getmtime(fpath) < cutoff:
            os.remove(fpath)
    print(f'[auto-backup] wrote {path} '
          f'({len(payload["centres"])} centres, {len(payload["share_links"])} share_links)')

def _auto_backup_loop():
    while True:
        try:
            _run_scheduled_backup()
        except Exception as e:
            print(f'[auto-backup] failed: {e}')
        time.sleep(_AUTO_BACKUP_INTERVAL_SECONDS)

threading.Thread(target=_auto_backup_loop, daemon=True).start()

@app.route('/admin/backup/auto')
def admin_backup_auto_list():
    os.makedirs(_AUTO_BACKUP_DIR, exist_ok=True)
    files = sorted(
        (f for f in os.listdir(_AUTO_BACKUP_DIR) if f.startswith('auto_backup_') and f.endswith('.json')),
        reverse=True
    )
    return jsonify({'backups': files})

@app.route('/admin/backup/auto/<path:filename>')
def admin_backup_auto_download(filename):
    filename = secure_filename(filename)
    if not filename.startswith('auto_backup_') or not filename.endswith('.json'):
        return jsonify({'error': 'Not found'}), 404
    fpath = os.path.join(_AUTO_BACKUP_DIR, filename)
    if not os.path.isfile(fpath):
        return jsonify({'error': 'Not found'}), 404
    return send_file(fpath, mimetype='application/json', as_attachment=True, download_name=filename)

@app.route('/admin/backup/import', methods=['POST'])
def admin_backup_import():
    f = request.files.get('backup')
    if not f:
        return jsonify({'error': 'No file uploaded'}), 400
    try:
        data = json.loads(f.read().decode())
    except Exception:
        return jsonify({'error': 'Invalid JSON file'}), 400

    counts = {}
    with get_db() as conn:
        # id is inserted explicitly (not auto-assigned) for every table below so
        # relationships stay intact — e.g. centre_images.centre_id must keep
        # pointing at the same centre it did before the backup was taken.
        counts['centres'] = _restore_rows(conn, data.get('centres', []), 'centres', [
            'id','name','address','city','about','space_type','brand','price_from','price_unit',
            'seat_type','open_hours','amenities','transport','website','map_url','coordinates',
            'why_recommend','source','created_at','hubble_id','hotdesk_price','has_coworking','min_desks',
        ])
        counts['centre_images'] = _restore_rows(conn, data.get('centre_images', []), 'centre_images', [
            'id','centre_id','filename','label','is_primary','sort_order',
        ])
        counts['proposals'] = _restore_rows(conn, data.get('proposals', []), 'proposals', [
            'id','title','template','client_name','client_company','client_email','client_location',
            'team_size','space_type','area_required','budget','duration','selected_centres',
            'manual_centres','status','pptx_filename','pdf_filename','created_at',
        ])
        counts['share_links'] = _restore_rows(conn, data.get('share_links', []), 'share_links', [
            'id','token','label','centre_ids','centre_names','created_at','created_by','client_email',
            'client_phone','canonical_ids','personalised_message','recommended_ids','is_test',
        ])
        counts['link_events'] = _restore_rows(conn, data.get('link_events', []), 'link_events', [
            'id','token','event_type','centre_id','centre_name','dwell_seconds','ip_hash',
            'user_agent','booking_date','booking_time','created_at',
        ])
    return jsonify({'ok': True, **{f'imported_{k}': v for k, v in counts.items()}})

def _restore_rows(conn, rows, table, columns):
    placeholders = ','.join('?' * len(columns))
    col_list = ','.join(columns)
    n = 0
    for row in rows:
        try:
            conn.execute(
                f'INSERT OR IGNORE INTO {table} ({col_list}) VALUES ({placeholders})',
                [row.get(c) for c in columns]
            )
            n += 1
        except Exception:
            pass
    return n


if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5001))
    app.run(debug=False, host='0.0.0.0', port=port)
